# -*- coding: utf-8 -*-
"""Excel das unidades do Vale do Ouro cujos compradores ja assinaram.

Tres abas:
  Compradores  uma linha por pessoa (o que o Lucas pediu)
  Unidades     uma linha por unidade, para somar valor sem contar em dobro
  Por imobiliaria  resumo com formulas sobre a aba Unidades

A aba de resumo aponta para Unidades, e nao para Compradores, de proposito: 43 das 113
unidades tem dois compradores (titular + conjuge), entao somar valor pela aba de pessoas
contaria a mesma unidade duas vezes.
"""
import json
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import sys
BASE = sys.argv[1] if len(sys.argv) > 1 else "."
dados = json.load(open(f"{BASE}/excel-compradores.json", encoding="utf-8"))

FONTE = "Arial"
TINTA = "1F2937"
CABECALHO_FILL = PatternFill("solid", fgColor="1F2937")
CABECALHO_FONTE = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
CORPO = Font(name=FONTE, size=10, color=TINTA)
TITULO = Font(name=FONTE, size=14, bold=True, color=TINTA)
SUB = Font(name=FONTE, size=9, color="6B7280")
DOURADO = Font(name=FONTE, size=10, bold=True, color="A07C3B")
BORDA_BAIXO = Border(bottom=Side(style="thin", color="D1D5DB"))
MOEDA = '"R$" #,##0;("R$" #,##0);-'
DATA_BR = "DD/MM/YYYY"

wb = Workbook()

def escreve_cabecalho(ws, colunas, linha=1):
    for i, (titulo, largura, _) in enumerate(colunas, start=1):
        c = ws.cell(row=linha, column=i, value=titulo)
        c.fill = CABECALHO_FILL
        c.font = CABECALHO_FONTE
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = largura
    ws.row_dimensions[linha].height = 28
    ws.freeze_panes = ws.cell(row=linha + 1, column=1)

def iso_para_data(s):
    if not s:
        return None
    a, m, d = str(s).split("-")
    return date(int(a), int(m), int(d))

# ── Aba 1: Compradores ───────────────────────────────────────────────────────
ws = wb.active
ws.title = "Compradores"
COLS = [
    ("Empreendimento", 16, "texto"),
    ("Unidade", 12, "texto"),
    ("Quadra", 9, "texto"),
    ("Lote", 8, "texto"),
    ("Comprador", 40, "texto"),
    ("CPF", 16, "texto"),
    ("Papel", 22, "texto"),
    ("E-mail", 34, "texto"),
    ("Imobiliária", 34, "texto"),
    ("Valor da unidade", 17, "moeda"),
    ("Enviado em", 13, "data"),
    ("Assinou em", 13, "data"),
    ("Dias até assinar", 15, "num"),
]
escreve_cabecalho(ws, COLS)

for i, d in enumerate(dados, start=2):
    valores = [
        d["empreendimento"], d["unidade"], d["quadra"], d["lote"], d["comprador"],
        d["cpf"], d["papel"], d["email"], d["imobiliaria"], d["valorUnidade"],
        iso_para_data(d["enviadoEm"]), iso_para_data(d["assinouEm"]), d["diasAteAssinar"],
    ]
    for j, v in enumerate(valores, start=1):
        c = ws.cell(row=i, column=j, value=v)
        c.font = CORPO
        c.border = BORDA_BAIXO
        tipo = COLS[j - 1][2]
        if tipo == "moeda":
            c.number_format = MOEDA
        elif tipo == "data":
            c.number_format = DATA_BR
        elif tipo == "num":
            c.alignment = Alignment(horizontal="right")

fim_compradores = len(dados) + 1
ws.auto_filter.ref = f"A1:M{fim_compradores}"

# ── Aba 2: Unidades ──────────────────────────────────────────────────────────
por_unidade = {}
for d in dados:
    por_unidade.setdefault(d["unidade"], []).append(d)

wu = wb.create_sheet("Unidades")
COLS_U = [
    ("Empreendimento", 16, "texto"),
    ("Unidade", 12, "texto"),
    ("Quadra", 9, "texto"),
    ("Lote", 8, "texto"),
    ("Compradores", 52, "texto"),
    ("Qtd. compradores", 16, "num"),
    ("Imobiliária", 34, "texto"),
    ("Valor da unidade", 17, "moeda"),
    ("Última assinatura", 16, "data"),
]
escreve_cabecalho(wu, COLS_U)

unidades = sorted(
    por_unidade.items(),
    key=lambda kv: (max(x["assinouEm"] or "" for x in kv[1]), kv[0]),
    reverse=True,
)
for i, (unidade, ls) in enumerate(unidades, start=2):
    p = ls[0]
    valores = [
        p["empreendimento"], unidade, p["quadra"], p["lote"],
        " · ".join(x["comprador"] for x in ls), len(ls), p["imobiliaria"],
        p["valorUnidade"], iso_para_data(max(x["assinouEm"] or "" for x in ls)),
    ]
    for j, v in enumerate(valores, start=1):
        c = wu.cell(row=i, column=j, value=v)
        c.font = CORPO
        c.border = BORDA_BAIXO
        tipo = COLS_U[j - 1][2]
        if tipo == "moeda":
            c.number_format = MOEDA
        elif tipo == "data":
            c.number_format = DATA_BR
        elif tipo == "num":
            c.alignment = Alignment(horizontal="right")

fim_unidades = len(unidades) + 1
wu.auto_filter.ref = f"A1:I{fim_unidades}"

linha_total = fim_unidades + 2
wu.cell(row=linha_total, column=7, value="Total").font = DOURADO
tv = wu.cell(row=linha_total, column=8, value=f"=SUM(H2:H{fim_unidades})")
tv.font = DOURADO
tv.number_format = MOEDA
tq = wu.cell(row=linha_total, column=6, value=f"=SUM(F2:F{fim_unidades})")
tq.font = DOURADO
tq.alignment = Alignment(horizontal="right")
wu.cell(row=linha_total, column=2, value=f"=COUNTA(B2:B{fim_unidades})").font = DOURADO

# ── Aba 3: Por imobiliária ───────────────────────────────────────────────────
wi = wb.create_sheet("Por imobiliária")
COLS_I = [
    ("Imobiliária", 40, "texto"),
    ("Unidades", 12, "num"),
    ("Compradores", 14, "num"),
    ("Valor somado", 18, "moeda"),
]
escreve_cabecalho(wi, COLS_I)

nomes = sorted({d["imobiliaria"] for d in dados})
# Ordena pelo que mais aparece, para a leitura começar pelo topo.
nomes.sort(key=lambda n: -len({d["unidade"] for d in dados if d["imobiliaria"] == n}))

for i, nome in enumerate(nomes, start=2):
    wi.cell(row=i, column=1, value=nome).font = CORPO
    # SUMIFS/COUNTIFS contra a aba Unidades: uma linha por unidade, sem dupla contagem.
    wi.cell(row=i, column=2, value=f"=COUNTIFS(Unidades!$G$2:$G${fim_unidades},$A{i})")
    wi.cell(row=i, column=3, value=f"=COUNTIFS(Compradores!$I$2:$I${fim_compradores},$A{i})")
    wi.cell(row=i, column=4, value=f"=SUMIFS(Unidades!$H$2:$H${fim_unidades},Unidades!$G$2:$G${fim_unidades},$A{i})")
    for j in range(2, 5):
        c = wi.cell(row=i, column=j)
        c.font = CORPO
        c.border = BORDA_BAIXO
        c.number_format = MOEDA if j == 4 else "#,##0"
        c.alignment = Alignment(horizontal="right")
    wi.cell(row=i, column=1).border = BORDA_BAIXO

fim_imob = len(nomes) + 1
lt = fim_imob + 2
wi.cell(row=lt, column=1, value="Total").font = DOURADO
for j, col in [(2, "B"), (3, "C"), (4, "D")]:
    c = wi.cell(row=lt, column=j, value=f"=SUM({col}2:{col}{fim_imob})")
    c.font = DOURADO
    c.number_format = MOEDA if j == 4 else "#,##0"
    c.alignment = Alignment(horizontal="right")

# ── Aba 4: Leia-me ───────────────────────────────────────────────────────────
wl = wb.create_sheet("Leia-me")
wl.column_dimensions["A"].width = 110
linhas_texto = [
    ("Unidades do Vale do Ouro com o comprador já assinado", TITULO),
    ("", CORPO),
    ("Extraído do C2X em 13/08/2026. Somente leitura, nada foi alterado no sistema.", CORPO),
    ("", CORPO),
    ("O que entra nesta lista", DOURADO),
    ("Unidades de VOC (Cecílio) e VOL (Lino) cujo contrato saiu para assinatura e ainda não foi", CORPO),
    ("cancelado, e em que TODOS os compradores daquela unidade já assinaram. Se o titular assinou", CORPO),
    ("e o cônjuge não, a unidade fica de fora.", CORPO),
    ("", CORPO),
    ("As três abas", DOURADO),
    ("Compradores: uma linha por pessoa. É onde estão CPF e e-mail.", CORPO),
    ("Unidades: uma linha por unidade, com os compradores juntos. Use esta para somar valor.", CORPO),
    ("Por imobiliária: resumo calculado por fórmula sobre a aba Unidades.", CORPO),
    ("", CORPO),
    ("Por que o valor é somado na aba Unidades e não na de Compradores", DOURADO),
    ("Porque 43 das 113 unidades têm dois compradores. Somando pela aba de pessoas, essas", CORPO),
    ("unidades entrariam duas vezes e o total ficaria inflado.", CORPO),
    ("", CORPO),
    ("De onde vem cada coisa", DOURADO),
    ("Comprador: quem assina com perfil Cliente no C2X, com o nome e o CPF congelados no envio.", CORPO),
    ("Imobiliária: o vínculo do cliente no C2X (users.vinculed_by_id). O campo de corretor da", CORPO),
    ("proposta está vazio em todas estas 113 unidades, por isso não foi usado.", CORPO),
    ("Assinou em: a data da assinatura na D4Sign. Quando há dois compradores, a aba Unidades", CORPO),
    ("mostra a data do último a assinar.", CORPO),
    ("Valor da unidade: o preço de tabela da unidade no C2X, não o valor do contrato.", CORPO),
    ("", CORPO),
    ("Celular não entrou: está vazio nos 156 registros.", SUB),
]
for i, (txt, fonte) in enumerate(linhas_texto, start=1):
    c = wl.cell(row=i, column=1, value=txt)
    c.font = fonte
    c.alignment = Alignment(vertical="center", wrap_text=False)

wb.move_sheet("Leia-me", offset=-3)

# Nao ha LibreOffice nesta maquina para pre-calcular, entao o openpyxl grava as formulas sem
# valor em cache. Sem esta linha, o Excel poderia mostrar celula vazia ate alguem forcar F9;
# com ela, ele recalcula tudo ao abrir o arquivo.
wb.calculation.fullCalcOnLoad = True

saida = f"{BASE}/compradores-assinados-vale-do-ouro.xlsx"
wb.save(saida)
print("gerado:", saida)
print(f"  Compradores: {len(dados)} linhas")
print(f"  Unidades: {len(unidades)} linhas")
print(f"  Por imobiliaria: {len(nomes)} linhas")
